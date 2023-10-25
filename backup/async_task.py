# -*- coding: UTF-8 -*-

import os
import shutil
import sys
import errno
from pathos.multiprocessing import ProcessingPool as Pool
from src.storage_connection import StorageConnection
import spacy
import re
import random
from datetime import datetime, timedelta
from dateutil.relativedelta import relativedelta
from collections import Counter
import pandas as pd
import time as timecheck
import itertools
from openpyxl import Workbook
import copy
from src import logger

_BASE_DIR = os.path.realpath(os.path.dirname(__file__))
sys.path.append(os.path.join(_BASE_DIR, '../../'))

TOTAL = 100
nlp = spacy.load('en_core_web_sm', disable=['textcat', 'parser', 'ner'])
_SAVE_DIR = '/var/tmp/tmi/history/results/'

es_index = "xvs-search-community"
src = "title,text,postdate"

query = {
    "query": {
        "bool": {
            "filter": [
                {
                    "match_phrase": {
                        "lang": "en"
                    }
                },
                {
                    "range": {
                        "timestamp": {
                            "format": "strict_date_optional_time"
                        }
                    }
                }
            ]
        }
    }
}

match_phrase = {
    "match_phrase": {
    }
}

def detection(tdate, category, domain, callback):
    try:
        y, m, d = map(int, tdate.split('-'))
        time = datetime(y, m, d)
        time_1 = time - timedelta(days=1)
        full_time = time - relativedelta(months=1)

        full_start = timecheck.time()

        # 1. t-1 ~ t 시간 데이터
        processor = Processor(time_1, 0, category, domain)
        tokens = processor.multi_processing()
        tokens = list(itertools.chain.from_iterable(tokens[0]))
        
        # 2. 전체 데이터 (한 달)
        processor = Processor(full_time, 1, category, domain)
        f_tokens = processor.multi_processing()
        f_tokens = list(itertools.chain.from_iterable(f_tokens[0]))
        
        tokens = count_tokens(tokens)
        f_tokens = count_tokens(f_tokens)

        full, K, n = full_data(tokens, f_tokens)
        new = new_words(n, K, tokens)
        re_emerging = re_emerging_words(full, tokens)
        logger.debug(f'[Word Detection full time]: {timecheck.time() - full_start}')

        excelpath = result2excel(new, re_emerging, tdate, category, domain)

        new = ', '.join(new)
        re_emerging = ', '.join(re_emerging)

        return new, re_emerging, excelpath

    except Exception as exception:
        logger.error(f'{exception}')


# data result to excel (openpyxl)
def result2excel(new, re_emerging, tdate, category, domain):
    filename = ""
    try:
        # excel 파일 생성
        wb = Workbook()
        for sheet in wb.sheetnames:
            # default sheet remove
            wb.remove(wb[sheet])

        nrsheet = wb.create_sheet(title='New & Re-emerging Words Detection', index=0)

        nrsheet = wb.active
        nrsheet.append(["[Input]"])
        nrsheet.append([f"Date: {tdate}"])
        nrsheet.append([f"Category: {category}"])
        nrsheet.append([f"Domain: {domain}"])
        nrsheet.append(['[New Words Result]'])
        for n in new:
            nrsheet.append([n])
        
        nrsheet.append([" "])
        nrsheet.append(["[Re-emerging Words Result]"])
        for re in re_emerging:
            nrsheet.append([re])
            
        filename = tdate + '_WordsDetection.xlsx'
        wb.save(_SAVE_DIR + filename)
        wb.close()
    except Exception as exception:
        logger.error(f'{exception}')

    return filename


class Processor():
    def __init__(self, time, flag, category, domain):
        self.pool = Pool(processes=8)
        self.time = time
        self.flag = flag
        self.category = category
        self.domain = domain

    # 1. Data Collection
    def data_collection(self, timestamp):
        es = StorageConnection.connect_elastic('kbn-prod')

        try:
            if self.flag == 0:
                timestamp_1 = (timestamp + timedelta(hours=3)).strftime('%Y-%m-%dT%H:%M:%S.%fZ')
                timestamp = timestamp.strftime('%Y-%m-%dT%H:%M:%S.%fZ')
            else:
                timestamp_1 = (timestamp + timedelta(days=4)).strftime('%Y-%m-%dT%H:%M:%S.%fZ')
                timestamp = timestamp.strftime('%Y-%m-%dT%H:%M:%S.%fZ')
            
            query1 = copy.deepcopy(query)
            query1['query']['bool']['filter'][1]['range']['timestamp']['gte'] = timestamp
            query1['query']['bool']['filter'][1]['range']['timestamp']['lte'] = timestamp_1
            
            category_cnt = copy.deepcopy(match_phrase)
            match_domain = copy.deepcopy(match_phrase)
            if self.category:
                category_cnt['match_phrase']['category_cnt'] = self.category
                query1['query']['bool']['filter'].append(category_cnt)
            if self.domain:
                match_domain['match_phrase']['domain'] = self.domain
                query1['query']['bool']['filter'].append(match_domain)
            
            count = es.count(index=es_index, body=query1)
            
            size = count['count']
            
            query2 = copy.deepcopy(query1)
            query2['size'] = int(size / 8)
            query2['sort'] = [{"_id": "asc"}, {"timestamp": "asc"}]
            
            community_index = es.search(index = es_index, _source = src, body = query2)
            
            bookmark = [community_index['hits']['hits'][-1]['sort'][0], community_index['hits']['hits'][-1]['sort'][1]]
            query3 = copy.deepcopy(query2)
            query3['search_after'] = bookmark
            
            result = []
            for html in map(lambda item: item['_source'], community_index['hits']['hits']):
                temp = f"{html['title']}\n{html['text']}"
                # Data Preprocessing
                result.extend([(token.lemma_.lower(), html['postdate']) for token in nlp(temp, disable=['textcat', 'parser', 'ner']) if not token.is_stop and token.pos_ in ('PROPN', 'NOUN', 'PRON')])

            hits_cnt = len(community_index['hits']['hits'])
                
            while hits_cnt < size:
                res = es.search(index = es_index, _source = src, body = query3)
                if res['hits']['hits']:
                    bookmark = [res['hits']['hits'][-1]['sort'][0], community_index['hits']['hits'][-1]['sort'][1]]
                    logger.debug(f'bookmark: {bookmark}')

                    for html in map(lambda item: item['_source'], res['hits']['hits']):
                        temp = f"{html['title']}\n{html['text']}"
                        # Data Preprocessing
                        result.extend([(token.lemma_.lower(), html['postdate']) for token in nlp(temp, disable=['textcat', 'parser', 'ner']) if not token.is_stop and token.pos_ in ('PROPN', 'NOUN', 'PRON')])
                
                    query3['size'] = int(size / 8)
                    query3['search_after'] = bookmark
                    hits_cnt += len(res['hits']['hits'])
                else:
                    return result
            return result
        except Exception as e:
            logger.error(e)

    def multi_processing(self):
        multi_time = []
        if self.flag == 0:
            time_split = [(self.time + timedelta(hours=3 * i)) for i in range(8)]
        else:
            time_split = [(self.time + timedelta(days=4 * i)) for i in range(8)]
        multi_time.append(self.pool.map(self.data_collection, time_split))

        return multi_time


# 2-1. Count tokens
def count_tokens(data):
    word = list(zip(*data))[0]
    count = Counter(word)
    duplicate = Counter([(dat[0], dat[1]) for dat in data])

    # 전체 단어 개수 - 같은 문장 내 같은 단어 개수 (2 이상이면 1만 남기고 빼기)
    for dup in duplicate:
        if duplicate[dup] >= 2:
            count[dup[0]] = count[dup[0]] - (duplicate[dup] - 1)

    count = dict(count)
    words = [[d[0], d[1], count.get(d[0])] for d in data]
    
    logger.debug(f'count tokens: {words[:3]}')
    return words


# 3. 사전
def load_dictionary(flag):
    dic_path = '/usr/src/tmi/src/wordsdetection/W2E/data/'
    Dtech = dic_path + 'D_tech.txt'

    # new = 0 / re-emerging = 1
    if flag == 0:
        # Dtech + Dcommon
        Dcommon = dic_path + 'D_common.txt'
        
        dic = []
        with open(Dcommon, 'r') as fc:
            temp = fc.read().split("\n")
            dic = dic + temp

        with open(Dtech, 'r') as ft:
            temp = ft.read().split("\n")
            dic = dic + temp
        return dic

    else:
        Dwhitelist = dic_path + 'D_whitelist.txt'

        with open(Dtech, 'r') as ft:
            tech = ft.read().split('\n')
        
        with open(Dwhitelist, 'r') as fw:
            whitelist = fw.read().split('\n')
        
        dic = [t for t in tech if t not in whitelist]
        return dic

# t-1 ~ t 시간 내의 단어 집합 구하기 & 키워드 집합 K 구하기
def full_data(tokens, full):

    # K: full 전체 단어 집합에서 90% 추출
    random.seed(100)
    K = [random.choice(full)[0] for _ in range(int(len(full) * 0.9))]

    # n (전체 기간 - t-1~t 사이 집합 K를 포함하는 문서 수) 구하기
    tmp_token = []
    tmp_time = ""
    words = []
    for idx, token in enumerate(tokens):
        # 같은 문서(같은 timestamp) 안에 있는 단어들 모으기 -> tmp_time
        if token[1] == tmp_time:
            words.append(token[0])
        else:
            if idx != 0:
                tmp_token.append(words)
                words = []
        tmp_time = token[1]
    
    n = 0
    # K 단어들을 포함하는 문서 개수 세기
    for tmp in tmp_token:
        for t in set(tmp):
            if t in K:
                n += 1
    logger.debug(f'len(n) = {n}')

    return full, K, n

    
# 4. New Words Detection
def new_words(n, K, C):
    func = lambda x, y : (x * y) / (x - y) >= 2.7
    dic = load_dictionary(0)
    dic_k = list(set(dic + K))

    # C_clean = C - dic - K = C - (dic + K)
    C_clean = [token for token in C if token[0] not in dic_k]
    # C_dic = [token for token in C if token[0] not in dic]
    # C_clean = [tok for tok in C_dic if tok[0] not in K]
    # C_clean = [token for token in C if token[0] not in (dic, K)]

    logger.debug(f'len(C) = {len(C_clean)}')    
    result = [word[0] for word in C_clean if func(n, int(word[2]))]

    # 중복 제거
    result = list(set(result))
    logger.debug(f'[New Words] : {result}')
    return result

# 5. Re-emerging Words Detection
def re_emerging_words(full, C):
    k = 15
    smoothing_f = 0.4
    sf = (smoothing_f * (1 - (1 - smoothing_f)**(2 * k))) / (2 - smoothing_f)
    result = []
    
    # timestamp column 기준으로 오름차순 정렬
    dfull = pd.DataFrame(full, columns=['word', 'timestamp', 'freq']).sort_values(by=['timestamp'], ascending=[True])
    # EWMA 계산
    dfull['EWMA'] = dfull['freq'].ewm(alpha=smoothing_f, adjust=False).mean()
    
    # (속도를 위해) 같은 단어 중복 제거
    dfull = dfull.drop_duplicates(['word'], keep='first')
    
    dic = load_dictionary(1)
    # CR = C 교집합 dic
    CR = [token for token in C if token[0] in dic]
    logger.debug(f'len(CR) = {len(CR)}')

    calc_list = []
    for index, cr in enumerate(CR):
        idx = dfull.index[dfull['word'] == cr[0]]
        try:
            ewma = float(dfull.iloc[idx]['EWMA'])
        except:
            # full 에 해당 단어가 존재하지 않는 경우
            continue
            
        calc_a = (float(cr[2]) - ewma)**2
        calc_list.append(calc_a)
        # 시그마 값 계산
        if index in (0, 1):
            continue
        calc = 3.8 * (sum(calc_list) / k) * sf

        del calc_list[0]
        if calc_a >= calc:
            result.append(cr)

    result = [r[0] for r in result]
    # 중복 제거
    result = list(set(result))
    logger.debug(f'[Re-emerging Words] : {result}')

    return result